"""
This file aims to calculate the zeroth and first order intensity as a result of
X-ray diffraction using phonon distribution data, along with the equations 
outlined in "X-Ray Thermal Diffuse Scattering And Its Studies of Lattice Dynamics" 
By Ruqing Xu, using data acquired by Lucas Lindsay.

Created July 9, 2015 by John Newman
"""
from math import atanh, tanh
import numpy as np
import cmath
from decimal import *
from openpyxl import *
#12295
# Loading Excel Workbook which contains Lucas' Data
wb = load_workbook('nonequilibrium diamond pure 300K.xlsx')
# Loading the Excel Worksheet which contains Lucas' Data
ws = wb.active
# y is the starting number of the cells to acquire and set data in the notebook
y = 8
# z is the ending number of the cells to acquire and set data in the notebook
# Note: due to the nature of range, we go to the last cell row + 1
z = 12296

def mag2(x):
	"""
	Helper function to calculate magnitude squared of a vector

	Paramater: x is a vector

	Returns: float
	"""
	return np.linalg.norm(x)


# Reduced Plank Constant with units (m**2 kg / s) 
h = 1.05457173 * 10**(-34)
#h = 1
# Angular Frequency of the jth phonon mode with wavevector k
# Note: Converts from THz to rad/s
w = range(y, z)
for i in range(y, z):
	w[i - y] = np.array(ws['F' + str(i)].value)*6.2831853e+12
print w
# Boltzmann's Constant with units (m**2 kg s**-2 K**-1)
kB = 1.3806488 * 10**(-23)
#kB = 1
# The value of the Absolute Temperature of the system, in units of K
T = 300
# eta_0 is the Bose-Einstein distribution
eta_0 = np.array([0.]*len(w))
for i in range(y, z):
		eta_0[i - y] = 1/((np.exp((h*w[i - y])/(kB*T))) - 1.)
		ws['K' + str(i)] = eta_0[i - y]
		print ws['K' + str(i)].value
# Temperature Gradient
Grad_T = np.array([1, 0, 0])
# F is a measure of how far the phonon is from equilibrium. Equation 10 in the
# paper provided by Lucas
F = range(y, z)
for i in range(y, z):
	F[i - y] = np.array(ws['H' + str(i)].value)
# eta_1 is the data which differs from the Bose-Einstein Data
print sum(np.dot(Grad_T, F[0]))
eta_1 = np.array([0.]*len(w))
for i in range(y, z):
	eta_1[i - y] = ((-eta_0[i-y]*(eta_0[i-y] + 1))/(kB*T))*sum(np.dot(Grad_T, \
		F[i - y]))
	ws['L' + str(i)] = eta_1[i - y]

# eta in total is the value which replaces the Bose-Einstein distribution in
# Xu's equations
eta = np.array([0.]*len(w))
for i in range(len(w)):
	eta[i] = eta_0[i] + eta_1[i]
# Number of unit cells
N = 100000
# Generalized coordinate squared
a2 = np.array([0.]*len(w))
for i in range(len(w)):
		a2[i] = (2*h/(N*w[i]))*(eta[i] + 1./2)
# Form factor of the s-th atom in the unit cell
f = [1., 1.1]
# I_incident is the photon flux of the incident x-ray beam in units 
# of photons/m^2
I_incident = 1
# Distance from the electron to the observer in units of m
d = 1
# The scattering length, also known as classical electron radius, in units of m
r_0 = 2.82 * 10**(-15)
# The scattering angle in units of rads
TwoTheta = np.pi/2
# The angle between the scattering plane and the plane defined by the
# polaization and the direction of the incident beam in units of rads
phi = np.pi/2
# Note: Should include diagram to provide visual aide for theta and phi
# Intensity measured for the Thomsom Scattering of a single electron for an
# x-ray that is linearly polarized
I_e = I_incident*(r_0**2/(d**2))*(np.sin(phi)**2 +\
 (np.cos(phi)**2)*np.cos(TwoTheta))
# Lattice Parameter of the Relaxed Diamond Lattice
a = 0.35668e-9
# Basis vectors of the Diamond Unit cell
tau = np.array([[ 0, 0, 0], [a/4., a/4., a/4.]])
# Scattering vector
q = []
for i in range(y, z):
	q.append(np.array([ws['C'+str(i)].value ,ws['D'+str(i)].value, \
		ws['E'+str(i)].value]))
# The mass of the s-th atom
# Note: Only one value since both diamond and germanium are monatomic
mu = 1
# Debye Waller Factor
M_s = np.array([0.]*len(w))
for i in range(y, z):
	M_s[i - y] = (eta[i - y] + 1./2)/w[i - y]
	M_s[i - y] = (h/(2*N*mu))*M_s[i - y]
	ws['N' + str(i)] = M_s[i - y]
M_s = sum(M_s)
# Reciprocal Lattice Vectors of the Unit Cell
k_l = np.array([[0.,0.,0.], [0.,2.,2.], [2.,0.,2.], [2.,2.,0.], \
	[3.,3.,3.], [3.,1.,1.], [1.,3.,1.], [1.,1.,3.]])*a/4.
# Zeroth Order Intensity (Bragg Diffraction)
I_0 = np.array([0.]*len(w))
for i in range(len(tau)):
	for j in range(len(k_l)):
		for l in range(y, z):
			I_0[l-y] = mag2(np.exp(-M_s[l-y])*f[i]*np.exp(-1j*\
				np.dot(k_l[j], tau[i])))
			I_0[l - y] = N*I_e*I_0[l - y]
			ws['O' + str(l)] = I_0[l - y]
# First Order Intensity (First Order Thermal Diffuse Scattering)
I_1 = np.array([0.]*len(w))
for i in range(y, z):
	for x in range(len(tau)):
		I_1[i-y] = ((eta[i-y] + 1./2)/w[i-y]) * \
			mag2((f[x]/mu)*np.exp(-M_s[i-y] - 1j*np.dot(q[i-y], tau[x])))
		I_1[i-y] = I_1[i-y]*N*I_e/2.
		ws['P' + str(i)] = I_1[i - y]
# Saves results in eta_results xl notebook
wb.save('eta_results.xlsx')
